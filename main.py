import openpyxl

def swap(list, pos1, pos2):
    list[pos1], list[pos2] = list[pos2], list[pos1]
    return list
def proeceesxl(mypath, your_skill, recommand):
    dataframe = openpyxl.load_workbook(mypath)

    dataframe._active_sheet_index = 0
    dataframe1 = dataframe.active
    maxcol = 6

    # init variant
    numbe = []
    Occup = []
    S_cat = []
    Skill = []
    virtu = []
    print_data = []
    print_final = []
    print_data_cnt = []
    index = 0
    print_occur = []
    # read data
    for col in dataframe1.iter_cols(1, maxcol):
        index += 1
        for row in range(1, dataframe1.max_row):
            if index == 1:
                numbe.append(col[row].value)
            elif index == 2:
                Occup.append(str(col[row].value))
            elif index == 3:
                S_cat.append(str(col[row].value))
            elif index == 4:
                Skill.append(str(col[row].value))
            elif index == 5:
                virtu.append(str(col[row].value))
            else:
                break

    your_skill = your_skill.lower()

    # main algorithm
    str_skill = ""
    ind_Skill = 0
    while ind_Skill < (len(Skill)):
        if Skill[ind_Skill].lower() == your_skill:
            str_skill = S_cat[ind_Skill].lower()
            break
        elif S_cat[ind_Skill].lower() == your_skill:
            str_skill = Occup[ind_Skill].lower()
            your_skill = str_skill
            ind_Skill =0
        ind_Skill+=1

    for ind_Skill in range(len(Skill)):
        if Skill[ind_Skill].lower() == your_skill:
            str_occup = Occup[ind_Skill].lower()
            if str_occup != "none":
                real_occu = -1
                for ind_occu in range(ind_Skill):
                    if Occup[ind_occu].lower() == str_occup:
                        real_occu = ind_occu
                        first_occr = ind_occu
                        print_occur.append(Occup[ind_occu])
                        break
                if real_occu != -1:
                    l_temp = []
                    while Occup[real_occu].lower() == str_occup:
                        if(S_cat[real_occu].lower() == str_skill and Skill[real_occu].lower() != "none"
                                and Skill[real_occu].lower() != your_skill):
                            l_temp.append(Skill[real_occu])
                        real_occu+=1
                    print_data.append(l_temp)

                        # break
        # Give priority to those that have higher number
    for i in range(len(print_data)):
        for j in range(len(print_data[i])):
            print_final.append(print_data[i][j])
    i = len(print_final) - 1
    print_data.clear()
    sum = 0
    for i in range(len(print_final)):
        sum =0
        for j in range(len(print_final)):
            if print_final[i] == print_final[j]:
                sum+=1
        print_data_cnt.append(sum)
    #  remove same
    i = len(print_final) - 1
    while i >= 0:
        flag = False
        for j in range (i):
            if print_final[i] == print_final[j]:
                print_final.pop(i)
                print_data_cnt.pop(i)
                break
        i -=1

    # Give priority to those that have higher number
    for i in range(len(print_final) -1):
        for j in range(i + 1, len(print_final)):
            if(print_data_cnt[i] < print_data_cnt[j]):
                swap(print_data_cnt, i, j)
                swap(print_final, i, j)
    # restrict the number of recommendations
    for i in range (int(recommand)):
        if (i > len(print_final) - 1 ):
            break
        print(print_final[i] + ", density number of co-occurence = " + str(print_data_cnt[i]))



# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    file_path=input("Enter data path: ")
    file_path.replace("\\", "/")
    your_skill = input('Enter your skill: ')
    recommand = input('Enter number of recommendations : ')
    if your_skill != "" and recommand != "":
        proeceesxl(file_path, your_skill, recommand)
        # proeceesxl('D:\\pythonproject\\data\\Sample Data 1 (2).xlsx', your_skill, recommand)
# See PyCharm help at https://www.jetbrains.com/help/pycharm/
